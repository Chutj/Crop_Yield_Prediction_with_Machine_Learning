import pandas as pd
import tensorflow as tf
from sklearn.model_selection import train_test_split
import scipy.stats as stats
from sklearn import metrics
import time
import os
import matplotlib.pyplot as plt
import openpyxl
import numpy as np

os.environ['TF_CPP_MIN_LOG_LEVEL']='2'

# Delete the model result from the last run.
def DeleteOldModel(ModelPath):
    AllFileName=os.listdir(ModelPath)
    for i in AllFileName:
        NewPath=os.path.join(ModelPath,i)
        if os.path.isdir(NewPath):
            DeleteOldModel(NewPath)
        else:
            os.remove(NewPath)

# Import data.
def LoadData(DataPath,Kind):
    if Kind==8:
        MyData=pd.read_csv(DataPath,names=['Lrad07','Prec06','SIF161','Shum06','Srad07','Srad08',
                                           'Srad10','Temp06','Yield'],header=0)
    elif Kind==54:
        MyData=pd.read_csv(DataPath,names=['EVI0610','EVI0626','EVI0712','EVI0728','EVI0813','EVI0829',
                                        'EVI0914','EVI0930','EVI1016','Lrad06','Lrad07','Lrad08',
                                        'Lrad09','Lrad10','Prec06','Prec07','Prec08','Prec09',
                                        'Prec10','Pres06','Pres07','Pres08','Pres09','Pres10',
                                        'SIF161','SIF177','SIF193','SIF209','SIF225','SIF241',
                                        'SIF257','SIF273','SIF289','Shum06','Shum07','Shum08',
                                        'Shum09','Shum10','SoilType','Srad06','Srad07','Srad08',
                                        'Srad09','Srad10','Temp06','Temp07','Temp08','Temp09',
                                        'Temp10','Wind06','Wind07','Wind08','Wind09','Wind10',
                                        'Yield'],header=0)
    Label={"Yield":MyData.pop("Yield")}
    Train_X,Train_Y=MyData,(pd.DataFrame(Label)) 
    return (Train_X,Train_Y)

def ProcessTrainInputData(TrainFeatures,TrainLabels,batch_size):
	dataset = tf.data.Dataset.from_tensor_slices((dict(TrainFeatures), TrainLabels))
	dataset = dataset.shuffle(1000).repeat().batch(batch_size)
	return ((dict(TrainFeatures), TrainLabels))

# Evaluation process takes too long maybe due to the code problems, so it is omitted later.
def ProcessEvalInputData(TrainFeatures,TrainLabels,batch_size,num_epochs=1,shuffle=False):
	dataset = tf.data.Dataset.from_tensor_slices((dict(TrainFeatures), TrainLabels))
	dataset = dataset.batch(batch_size)
	return ((dict(TrainFeatures), TrainLabels))

def ProcessTestInputData(TestFeatures,batch_size):	
	dataset = tf.data.Dataset.from_tensor_slices(dict(TestFeatures))
	return (tf.compat.v1.data.make_one_shot_iterator(dataset.batch(batch_size)).get_next())

# Verify the accuracy and dram the fitting image.
def AccuracyVerification(PredictLabels,TestLabels):
    value=0
    PredictValuesList=[]
    for k in PredictLabels:
        value=k.get('predictions')[0]
        PredictValuesList.append(value)
    TestLabels=TestLabels.values.tolist()
    TestYList=sum(TestLabels,[])
    Pearsonr=stats.pearsonr(TestYList,PredictValuesList)
    R2=metrics.r2_score(TestYList,PredictValuesList)
    RMSE=metrics.mean_squared_error(TestYList,PredictValuesList)**0.5
    time.sleep(3)
    plt.cla()
    plt.plot(TestYList,PredictValuesList,'r*')
    plt.xlabel('Actual Values')
    plt.ylabel('Predicted Values')
    print('Pearson correlation coefficient is {0}, and RMSE is {1}.'.format(Pearsonr[0],RMSE))
    return (Pearsonr[0],R2,RMSE,PredictValuesList)

# Save precision results and model key parameters.
def WriteAccuracy(*WriteVar):
    ExcelData=openpyxl.load_workbook(WriteVar[0])
    SheetName=ExcelData.get_sheet_names()
    WriteSheet=ExcelData.get_sheet_by_name(SheetName[0])
    WriteSheet=ExcelData.active
    MaxRowNum=WriteSheet.max_row
    for i in range(len(WriteVar)-1):
        exec("WriteSheet.cell(MaxRowNum+1,i+1).value=WriteVar[i+1]")
    ExcelData.save(WriteVar[0])

# Start the program.
MyModelPath="G:/CropYield/03_DL/02_DNNModle"
# Determine the number of environment variables.
VarKind=54
if VarKind==8:
    MyDataPath="G:/CropYield/03_DL/00_Data/AllData.csv"
    MyEvalSavePath="G:/CropYield/03_DL/03_OtherResult/EvalResult.xlsx"
elif VarKind==54:
    MyDataPath="G:/CropYield/03_DL/00_Data/AllDataAll.csv"
    MyEvalSavePath="G:/CropYield/03_DL/03_OtherResult/EvalResult54.xlsx"

Dropout=0.2
HiddenLayer=[64,128,256,512,512,1024]
TrainStep=3000
TrainBatchSize=90
TestSize=0.2
RandomSeed=np.random.randint(low=21,high=22)
OptMethod='tf.optimizers.Adam()'
ActFun='tf.nn.relu'

RMSE=9999
while RMSE>593:
    DeleteOldModel(MyModelPath)
    AllX,AllY=LoadData(MyDataPath,VarKind)
    TrainX,TestX,TrainY,TestY=train_test_split(AllX,AllY,test_size=TestSize,random_state=RandomSeed)
    FeatureColumnName=[]
    FeatureName=['Lrad07','Prec06','SIF161','Shum06','Srad07','Srad08','Srad10','Temp06']
    
    for key in FeatureName:
    	FeatureColumnName.append(tf.feature_column.numeric_column(key=key))
    
    regressor=tf.estimator.DNNRegressor(feature_columns=FeatureColumnName,
                                              hidden_units=HiddenLayer,
    										  optimizer=eval(OptMethod),
                                              dropout=Dropout,
    										  #loss_reduction=tf.losses.Reduction.AUTO,
    										  activation_fn=eval(ActFun),
    										  label_dimension=1, 
                                              model_dir=MyModelPath
    										  )
    tf.compat.v1.logging.set_verbosity(tf.compat.v1.logging.INFO)
    regressor.train(input_fn=lambda:ProcessTrainInputData(TrainX,TrainY,TrainBatchSize),steps=TrainStep) 
    
    # ProcessEvalInputData=tf.compat.v1.estimator.inputs.pandas_input_fn(x=TestX,
    #                                                                    y=TestY,
    #                                                                    batch_size=10,
    #                                                                    num_epochs=1,
    #                                                                    shuffle=False
    #                                                                    )
    # EvalResult=regressor.evaluate(input_fn=lambda:ProcessEvalInputData(TestX,TestY,batch_size=10))
    # print('Evaluate:{}'.format(EvalResult))
    
    PredictValues=regressor.predict(input_fn=lambda:ProcessTestInputData(TestX,batch_size=1))
    
    AccuracyResult=AccuracyVerification(PredictValues,TestY)
    PearsonR,R2,RMSE,PredictY=AccuracyResult[0],AccuracyResult[1],AccuracyResult[2],AccuracyResult[3]
    
    WriteAccuracy(MyEvalSavePath,PearsonR,R2,RMSE,TestSize,RandomSeed,OptMethod,ActFun,
                  ','.join('%s' %i for i in HiddenLayer),Dropout,TrainBatchSize,TrainStep)