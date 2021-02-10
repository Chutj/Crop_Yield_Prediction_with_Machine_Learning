import os
import openpyxl
import numpy as np
import pandas as pd
import seaborn as sns
import tensorflow as tf
import scipy.stats as stats
import matplotlib.pyplot as plt
from sklearn import metrics
from tensorflow import keras
from tensorflow.keras import layers
from tensorflow.keras import regularizers
from tensorflow.keras.callbacks import ModelCheckpoint
from tensorflow.keras.layers.experimental import preprocessing

np.set_printoptions(precision=4,suppress=True)

# Draw the joint distribution image.
def JointDistribution(Factors):
    plt.figure(1)
    sns.pairplot(TrainData[Factors],kind='reg',diag_kind='kde')
    sns.set(font_scale=2.0)
    DataDistribution=TrainData.describe().transpose()

# Delete the model result from the last run.
def DeleteOldModel(ModelPath):
    AllFileName=os.listdir(ModelPath)
    for i in AllFileName:
        NewPath=os.path.join(ModelPath,i)
        if os.path.isdir(NewPath):
            DeleteOldModel(NewPath)
        else:
            os.remove(NewPath)

# Find and save optimal epoch.
def CheckPoint(Name):
    Checkpoint=ModelCheckpoint(Name,
                               monitor=CheckPointMethod,
                               verbose=1,
                               save_best_only=True,
                               mode='auto')
    CallBackList=[Checkpoint]
    return CallBackList

# Build DNN model.
def BuildModel(Norm):
    Model=keras.Sequential([Norm,
                            
                            layers.Dense(HiddenLayer[0],
                                         kernel_regularizer=regularizers.l2(RegularizationFactor),
                                         activation=ActivationMethod),
                            layers.BatchNormalization(),
                            layers.Dropout(DropoutValue[0]),
                            
                            layers.Dense(HiddenLayer[1],
                                         kernel_regularizer=regularizers.l2(RegularizationFactor),
                                         activation=ActivationMethod),
                            layers.BatchNormalization(),
                            layers.Dropout(DropoutValue[1]),
                            
                            layers.Dense(HiddenLayer[2],
                                         kernel_regularizer=regularizers.l2(RegularizationFactor),
                                         activation=ActivationMethod),
                            layers.BatchNormalization(),
                            layers.Dropout(DropoutValue[2]),
                            
                            layers.Dense(HiddenLayer[3],
                                         kernel_regularizer=regularizers.l2(RegularizationFactor),
                                         activation=ActivationMethod),
                            layers.BatchNormalization(),
                            layers.Dropout(DropoutValue[3]),
                            
                            layers.Dense(HiddenLayer[4],
                                         kernel_regularizer=regularizers.l2(RegularizationFactor),
                                         activation=ActivationMethod),
                            layers.BatchNormalization(),
                            layers.Dropout(DropoutValue[4]),
                            
                            layers.Dense(HiddenLayer[5],
                                         kernel_regularizer=regularizers.l2(RegularizationFactor),
                                         activation=ActivationMethod),
                            layers.BatchNormalization(),
                            layers.Dropout(DropoutValue[5]),
                            
                            layers.Dense(HiddenLayer[6],
                                         kernel_regularizer=regularizers.l2(RegularizationFactor),
                                         activation=ActivationMethod),
                            # If batch normalization is set in the last hidden layer, the error image
                            # will show a trend of first stable and then decline; otherwise, it will 
                            # decline and then stable.
                            # layers.BatchNormalization(),
                            layers.Dropout(DropoutValue[6]),
                            
                            layers.Dense(1)])
    Model.compile(loss=LossMethod,
                  optimizer=tf.keras.optimizers.Adam(learning_rate=LearnRate,decay=LearnDecay))
    return Model

# Draw error image.
def LossPlot(History):
    plt.figure(2)
    plt.plot(History.history['loss'],label='loss')
    plt.plot(History.history['val_loss'],label='val_loss')
    plt.ylim([0,4000])
    plt.xlabel('Epoch')
    plt.ylabel('Error')
    plt.legend()
    plt.grid(True)

# Draw Test image.
def TestPlot(TestY,TestPrediction):
    plt.figure(3)
    ax=plt.axes(aspect='equal')
    plt.scatter(TestY,TestPrediction)
    plt.xlabel('True Values')
    plt.ylabel('Predictions')
    Lims=[0,13000]
    plt.xlim(Lims)
    plt.ylim(Lims)
    plt.plot(Lims,Lims)
    plt.grid(False)

# Verify the accuracy and draw error hist image.
def AccuracyVerification(TestY,TestPrediction):
    DNNError=TestPrediction-TestY
    plt.figure(4)
    plt.hist(DNNError,bins=30)
    plt.xlabel('Prediction Error')
    plt.ylabel('Count')
    plt.grid(False)
    Pearsonr=stats.pearsonr(TestY,TestPrediction)
    R2=metrics.r2_score(TestY,TestPrediction)
    RMSE=metrics.mean_squared_error(TestY,TestPrediction)**0.5
    print('Pearson correlation coefficient is {0}, and RMSE is {1}.'.format(Pearsonr[0],RMSE))
    return (Pearsonr[0],R2,RMSE)

# Save key parameters.
def WriteAccuracy(*WriteVar):
    ExcelData=openpyxl.load_workbook(WriteVar[0])
    SheetName=ExcelData.get_sheet_names()
    WriteSheet=ExcelData.get_sheet_by_name(SheetName[0])
    WriteSheet=ExcelData.active
    MaxRowNum=WriteSheet.max_row
    for i in range(len(WriteVar)-1):
        exec("WriteSheet.cell(MaxRowNum+1,i+1).value=WriteVar[i+1]")
    ExcelData.save(WriteVar[0])

# Input parameters.
DataPath="G:/CropYield/03_DL/00_Data/AllDataAll.csv"
ModelPath="G:/CropYield/03_DL/02_DNNModle"
CheckPointName="G:/CropYield/03_DL/02_DNNModle/Weights/Weights_{epoch:03d}_{val_loss:.4f}.hdf5"
ParameterPath="G:/CropYield/03_DL/03_OtherResult/ParameterResult.xlsx"
TrainFrac=0.8
RandomSeed=np.random.randint(low=21,high=22)
CheckPointMethod='val_loss'
HiddenLayer=[64,128,256,512,512,1024,1024]
RegularizationFactor=0.0001
ActivationMethod='relu'
DropoutValue=[0.5,0.5,0.5,0.3,0.3,0.3,0.2]
LossMethod='mean_absolute_error'
LearnRate=0.005
LearnDecay=5e-4
FitEpoch=200
ValFrac=0.2

# Fetch and divide data.
MyData=pd.read_csv(DataPath,names=['EVI0610','EVI0626','EVI0712','EVI0728','EVI0813','EVI0829',
                                   'EVI0914','EVI0930','EVI1016','Lrad06','Lrad07','Lrad08',
                                   'Lrad09','Lrad10','Prec06','Prec07','Prec08','Prec09',
                                   'Prec10','Pres06','Pres07','Pres08','Pres09','Pres10',
                                   'SIF161','SIF177','SIF193','SIF209','SIF225','SIF241',
                                   'SIF257','SIF273','SIF289','Shum06','Shum07','Shum08',
                                   'Shum09','Shum10','SoilType','Srad06','Srad07','Srad08',
                                   'Srad09','Srad10','Temp06','Temp07','Temp08','Temp09',
                                   'Temp10','Wind06','Wind07','Wind08','Wind09','Wind10',
                                   'Yield'],header=0)
TrainData=MyData.sample(frac=TrainFrac,random_state=RandomSeed)
TestData=MyData.drop(TrainData.index)

# Draw the joint distribution image.
# JointFactor=['Lrad07','Prec06','SIF161','Shum06','Srad07','Srad08','Srad10','Temp06','Yield']
# JointDistribution(JointFactor)

# Separate independent and dependent variables.
TrainX=TrainData.copy(deep=True)
TestX=TestData.copy(deep=True)
TrainY=TrainX.pop('Yield')
TestY=TestX.pop('Yield')

# Standardization data.
Normalizer=preprocessing.Normalization()
Normalizer.adapt(np.array(TrainX))

# Delete the model result from the last run.
DeleteOldModel(ModelPath)

# Find and save optimal epoch.
CallBack=CheckPoint(CheckPointName)

# Build DNN regression model.
DNNModel=BuildModel(Normalizer)
DNNModel.summary()
DNNHistory=DNNModel.fit(TrainX,
                        TrainY,
                        epochs=FitEpoch,
                        verbose=1,
                        callbacks=CallBack,
                        validation_split=ValFrac)

# Draw error image.
LossPlot(DNNHistory)

# Predict test set data.
TestPrediction=DNNModel.predict(TestX).flatten()

# Draw Test image.
TestPlot(TestY,TestPrediction)

# Verify the accuracy and draw error hist image.
AccuracyResult=AccuracyVerification(TestY,TestPrediction)
PearsonR,R2,RMSE=AccuracyResult[0],AccuracyResult[1],AccuracyResult[2]

# Save model and key parameters.
DNNModel.save(ModelPath)
WriteAccuracy(ParameterPath,PearsonR,R2,RMSE,TrainFrac,RandomSeed,CheckPointMethod,
              ','.join('%s' %i for i in HiddenLayer),RegularizationFactor,
              ActivationMethod,','.join('%s' %i for i in DropoutValue),
              LossMethod,LearnRate,LearnDecay,FitEpoch,ValFrac)
